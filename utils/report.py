import openpyxl

from database.models import Student, Attendance


def create_report(filename, students: list[Student]):
    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()

    # Получаем список всех уникальных групп
    groups = set([student.group_name for student in students])

    # Проходим по каждой группе и создаем соответствующий лист
    for group_name in groups:
        # Добавляем новый лист в книгу и получаем объект листа
        sheet = wb.create_sheet(title=group_name)

        # Получаем список всех студентов в группе
        group_students = [
            student for student in students if student.group_name == group_name
        ]

        # Заполняем первый столбец листа именами студентов
        for i, student in enumerate(group_students):
            sheet.cell(row=i + 1, column=1, value=student.name)

        # Получаем список всех дат, за которые есть посещения
        dates = sorted(
            set(
                [
                    attendance.date
                    for student in group_students
                    for attendance in student.attendances
                ]
            )
        )

        # Заполняем заголовки столбцов датами
        for i, date in enumerate(dates):
            sheet.cell(row=1, column=i + 2, value=date.strftime("%d.%m.%Y"))

        # Заполняем таблицу посещений
        for i, student in enumerate(group_students):
            for j, date in enumerate(dates):
                attendance = [
                    attendance
                    for attendance in student.attendances
                    if attendance.date == date
                ]
                if attendance:
                    sheet.cell(row=i + 2, column=j + 2, value="+")
                else:
                    sheet.cell(row=i + 2, column=j + 2, value="-")

    # Сохраняем файл
    wb.save(filename)
